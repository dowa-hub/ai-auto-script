"""
Unified script ingestion.

parse_unified(content, filename) does ONE pass through the document and
returns both the word list (for STT tracking) AND the display HTML in a
single result.  Because they're built together, line_index N in the word
list always corresponds to data-line="N" in the HTML — no drift.

PDF / image files return html=None (display is handled natively by the
frontend: PDF.js for PDFs, <img> for images).
"""
import io
import re
import html as _html
from pathlib import Path


# ── Public API ────────────────────────────────────────────────────────────────

def parse_unified(content: bytes, filename: str) -> dict:
    """
    Single-pass parse.  Returns:
      lines       – list of plain-text lines (one per script row/paragraph)
      words       – list of {word, clean, line_index}
      word_count  – int
      line_count  – int
      html        – HTML string with data-line="N" on every block element,
                    or None for PDF / image files
    """
    ext = Path(filename).suffix.lower()

    if ext == ".pdf":
        return _pdf_parse(content)
    if ext in (".docx", ".doc"):
        return _docx_parse(content)
    if ext in (".xlsx", ".xls"):
        return _excel_parse(content)
    if ext in (".txt", ".text"):
        return _txt_parse(content)
    if ext in (".jpg", ".jpeg", ".png", ".tiff", ".tif", ".bmp", ".webp"):
        return _image_parse(content)
    # Unknown — try plain text
    return _txt_parse(content)


# Keep parse_script as a backwards-compatible alias used by main.py
def parse_script(content: bytes, filename: str) -> dict:
    result = parse_unified(content, filename)
    return result


# ── Per-format parsers (each does one pass, builds words + HTML together) ─────

def _docx_parse(content: bytes) -> dict:
    from docx import Document
    lines, words, html_rows = [], [], []

    doc = Document(io.BytesIO(content))
    for child in doc.element.body:
        tag = child.tag.split("}")[-1] if "}" in child.tag else child.tag

        if tag == "p":
            text = "".join(
                r.text or "" for r in child.iter() if r.tag.split("}")[-1] == "t"
            )
            if not text.strip():
                continue
            line_idx = _add_line(lines, words, text)
            html_rows.append(f'<p data-line="{line_idx}">{_h(text)}</p>')

        elif tag == "tbl":
            html_rows.append("<table>")
            for tr in child.iter():
                if tr.tag.split("}")[-1] != "tr":
                    continue
                cells, seen = [], set()
                for tc in tr.iter():
                    if tc.tag.split("}")[-1] != "tc":
                        continue
                    ct = "".join(
                        r.text or "" for r in tc.iter() if r.tag.split("}")[-1] == "t"
                    ).strip()
                    if ct not in seen:
                        cells.append(ct)
                        seen.add(ct)
                row_text = "\t".join(cells)
                if not row_text.strip():
                    continue
                line_idx = _add_line(lines, words, row_text)
                tds = "".join(f"<td>{_h(c)}</td>" for c in cells)
                html_rows.append(f'<tr data-line="{line_idx}">{tds}</tr>')
            html_rows.append("</table>")

    return _result(lines, words, "\n".join(html_rows))


def _excel_parse(content: bytes) -> dict:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    lines, words, html_rows = [], [], []

    for ws in wb.worksheets:
        html_rows.append(f'<h3 class="sheet-name">{_h(ws.title)}</h3><table>')
        for row in ws.iter_rows(values_only=True):
            # Keep cells that are not None; skip rows that are entirely empty
            cells = [c for c in row if c is not None]
            row_text = "\t".join(str(c) for c in cells if str(c).strip())
            if not row_text.strip():
                continue
            line_idx = _add_line(lines, words, row_text)
            tds = "".join(f"<td>{_h(str(c))}</td>" for c in cells)
            html_rows.append(f'<tr data-line="{line_idx}">{tds}</tr>')
        html_rows.append("</table>")

    return _result(lines, words, "\n".join(html_rows))


def _txt_parse(content: bytes) -> dict:
    lines, words, html_rows = [], [], []
    for raw in content.decode("utf-8", errors="replace").split("\n"):
        if not raw.strip():
            continue
        line_idx = _add_line(lines, words, raw)
        html_rows.append(f'<p data-line="{line_idx}">{_h(raw)}</p>')
    return _result(lines, words, "\n".join(html_rows))


def _pdf_parse(content: bytes) -> dict:
    try:
        import pypdfium2 as pdfium
        pdf = pdfium.PdfDocument(content)

        # Detect "Script" column boundaries for table-based PDFs
        script_col = _detect_script_column(pdf)

        script_pages = []   # column-clipped (for word tracking)
        full_pages   = []   # full page text (for section detection with timestamps)
        for page in pdf:
            textpage = page.get_textpage()
            if script_col:
                left, right = script_col
                sp = textpage.get_text_bounded(
                    left=left, bottom=0,
                    right=right, top=page.get_height(),
                )
                fp = textpage.get_text_bounded()  # includes Time + Cues columns
            else:
                sp = textpage.get_text_bounded()
                fp = sp  # same — no column stripping
            script_pages.append(sp)
            full_pages.append(fp)

        if len("\n".join(script_pages).strip()) < 100:
            # OCR fallback — treat entire result as a single "page"
            script_pages = [_ocr_pdf(content)]
            full_pages   = script_pages
    except Exception as e:
        raise ValueError(f"PDF parse failed: {e}")

    # Build word list from script column only (accurate for STT tracking)
    lines, words = [], []
    line_pages: dict[int, int] = {}  # line_index → 0-based page number

    for page_num, page_text in enumerate(script_pages):
        for raw in page_text.split("\n"):
            if raw.strip():
                line_idx = _add_line(lines, words, raw)
                line_pages[line_idx] = page_num

    result = _result(lines, words, None)  # HTML=None → PDF.js handles display
    result["line_pages"] = line_pages

    # For table-format PDFs, detect sections by scanning the Time column
    # for timestamps and matching them to Script column content at the same
    # vertical position. This is structural (not text-heuristic) so it handles
    # any Script column content regardless of starting character (#, ●, etc.)
    if script_col:
        cell_sections = _detect_table_cells(pdf, script_col, words, line_pages, lines)
        result["sections"] = cell_sections if cell_sections else _detect_sections(lines, words)

    return result


def _detect_table_cells(pdf, script_col, script_words, line_pages, script_lines):
    """Detect sections from table-format PDFs by y-position matching.

    Scans the Time column for timestamps and pairs each with the Script column
    text at the same vertical position on the page. Structural approach — no
    text-pattern heuristics, works for any Script content (#1 Video, ●, etc.).

    Pages without any timestamp in the Time column are skipped (they are
    continuations of the previous section, not new sections).
    """
    left_x, right_x = script_col

    _TS_ANY = re.compile(
        r'(\d{1,2}:\d{2}\s*[ap]m|\d{1,2}:\d{2}|\d{1,2}[ap]m)',
        re.IGNORECASE,
    )
    # Skip page-header / column-header rows that repeat on every page
    _SKIP_RE = re.compile(
        r'^(Time\s*$|Script\s*$|Virtual|Cue|NOTE\b|AV\s*Team|Projection)',
        re.IGNORECASE,
    )
    # Stage-note content after timestamps — NOT real sections
    _STAGE_NOTE = re.compile(r'^(?:From\s+\w|Notes?:|TBD\b|N/A\b)', re.IGNORECASE)

    # Per-page script word lists in document order
    page_word_lists = {}
    for wi, w in enumerate(script_words):
        pg = line_pages.get(w["line_index"], 0)
        page_word_lists.setdefault(pg, []).append((wi, w))

    def _table_y_bounds(page):
        """Return (y_min, y_max) of the table content area using drawn H-lines."""
        hs = []
        for obj in page.get_objects():
            if obj.type == 2:
                l, b, r, t = obj.get_pos()
                if (t - b) < 3 and (r - l) > 200:
                    hs.append(b)
        if len(hs) >= 2:
            return min(hs) + 2, max(hs) - 2
        return 50, page.get_height() - 50

    def _merge_char_rows(char_map, gap=20):
        """Merge y-adjacent character groups into text rows [(y_top, text)]."""
        rows = []
        group_y, group_chars = None, []
        for y in sorted(char_map.keys(), reverse=True):
            if group_y is None or group_y - y > gap:
                if group_chars:
                    text = "".join(c for _, c in sorted(group_chars, key=lambda t: t[0]))
                    rows.append((group_y, text.strip()))
                group_y = y
                group_chars = list(char_map[y])
            else:
                group_chars.extend(char_map[y])
        if group_chars:
            text = "".join(c for _, c in sorted(group_chars, key=lambda t: t[0]))
            rows.append((group_y, text.strip()))
        return rows

    def _find_word_idx(pg_wds, script_text, used):
        """First unused page word matching first two alpha words of script_text."""
        alpha = re.findall(r"[A-Za-z]{2,}", script_text)
        key0 = alpha[0].lower() if alpha else ""
        key1 = alpha[1].lower() if len(alpha) > 1 else None

        def _m(clean, key):
            return clean == key or (len(key) >= 3 and clean.startswith(key[:min(len(key), 4)]))

        for i, (wi, w) in enumerate(pg_wds):
            if wi in used or not key0 or not _m(w["clean"], key0):
                continue
            if key1:
                nearby = [pw["clean"] for _, pw in pg_wds[i+1:i+5]]
                # key1 may be embedded in the matched word (e.g. "postevent" contains "event")
                if key1 not in w["clean"] and not any(_m(c, key1) for c in nearby):
                    continue
            return wi, w["line_index"]
        for wi, w in pg_wds:  # fallback: first unused
            if wi not in used:
                return wi, w["line_index"]
        return None, None

    def _make_summary(start_wi):
        """First ~7 speech words after start_wi, skipping speaker-cue lines."""
        _SP = re.compile(r'^[A-Z][a-zA-Z\-]+(?: [A-Z][a-zA-Z\-]+){0,3}:\s*$')
        i = start_wi
        while i < min(start_wi + 80, len(script_words)):
            li = script_words[i]["line_index"]
            line_wds = []
            j = i
            while j < len(script_words) and script_words[j]["line_index"] == li:
                line_wds.append(script_words[j]["word"])
                j += 1
            line_text = " ".join(line_wds).strip()
            # Strip the same column-bleed prefixes that title stripping handles
            line_text = re.sub(r'^(?:PM?|AM|MC)\s+', '', line_text)
            line_text = re.sub(r'^[A-Z]\s*[–—\-]\s*', '', line_text)  # "H — Gathering" → "Gathering"
            line_text = re.sub(r'^[A-Z]\s+', '', line_text)            # "M Poison Waters" → "Poison Waters"
            line_text = re.sub(r'^[A-Z](?=[A-Z][a-z])', '', line_text) # "PEvent" → "Event"
            # Skip lines starting with lowercase — truncated page headers or mid-sentence
            # fragments (e.g. "ebrook EQUIVOX", "ational Summit")
            if line_text and line_text[0].islower():
                i = j
                continue
            # Skip column headers and AV-team noise (e.g. "Script", "Virtual Cues")
            if line_text and _SKIP_RE.match(line_text):
                i = j
                continue
            if line_text and not _SP.match(line_text):
                snippet = line_text.split()[:7]
                return " ".join(snippet) + ("…" if len(snippet) >= 7 else "")
            i = j
        return ""

    sections = []
    used_wis = set()

    for pg_idx, page in enumerate(pdf):
        tp = page.get_textpage()
        n = tp.count_chars()
        if n == 0:
            continue

        y_min, y_max = _table_y_bounds(page)

        # Scan chars inside the table area only (excludes running page headers)
        # Time column: x < left_x-3  (small margin avoids "M" from "PM" leaking)
        # Script column: left_x+3 ≤ x ≤ right_x
        # Scan Time column chars only (for timestamp y-position detection)
        time_by_y = {}

        for i in range(min(n, 12000)):
            c = tp.get_text_range(i, 1)
            if not c or not c.strip():
                continue
            try:
                box = tp.get_charbox(i)
                x = box[0]
                y = (box[1] + box[3]) / 2
            except Exception:
                continue
            if not (y_min < y < y_max):
                continue          # outside table area — skip
            if x < left_x - 3:
                y_key = round(y / 8) * 8
                time_by_y.setdefault(y_key, []).append((x, c))

        time_rows = _merge_char_rows(time_by_y, gap=20)

        # Find timestamps in Time column rows
        ts_positions = []
        for y, text in time_rows:
            if _SKIP_RE.match(text):
                continue
            m = _TS_ANY.search(text)
            if m:
                ts_positions.append((y, m.group(1).strip()))

        if not ts_positions:
            continue  # pure continuation page — no new sections here

        pg_wds = page_word_lists.get(pg_idx, [])
        ts_positions.sort(key=lambda t: t[0], reverse=True)  # top → bottom

        for ts_idx, (ts_y, ts_str) in enumerate(ts_positions):
            # y-window for this cell: from ts_y down to next timestamp (or ts_y - 300)
            y_floor = (ts_positions[ts_idx + 1][0]
                       if ts_idx + 1 < len(ts_positions) else ts_y - 300)

            # Get Script column text in this cell's y-window using get_text_bounded.
            # This gives properly-spaced text (unlike char-by-char concatenation).
            cell_bottom = max(y_floor - 5, y_min)
            cell_top    = min(ts_y + 25, y_max)
            raw_cell = tp.get_text_bounded(
                left=left_x + 3, right=right_x,
                bottom=cell_bottom, top=cell_top,
            )
            first_script = ""
            for ln in raw_cell.replace('\r\n', '\n').replace('\r', '\n').split('\n'):
                ln = ln.strip()
                # Skip boundary-artifact chars and header rows.
                # Require at least one 4+ consecutive alpha sequence — filters "y ggg", "gpg".
                if ln and re.search(r'[A-Za-z]{4,}', ln) and not _SKIP_RE.match(ln):
                    first_script = ln
                    break

            # Strip leading column-indicator prefix for display and filtering.
            # Handles: "M Poison Waters:" → "Poison Waters:"    (single-letter + space)
            #          "PM Event Ends"   → "Event Ends"         (2-char + space)
            #          "PDoors Open"     → "Doors Open"         (no-space bleed from PM)
            stripped_script = re.sub(r'^(?:PM?|AM|MC)\s+', '', first_script)
            stripped_script = re.sub(r'^[A-Z]\s+', '', stripped_script)
            stripped_script = re.sub(r'^[A-Z](?=[A-Z][a-z])', '', stripped_script)

            # Skip stage notes appearing as timestamped rows (e.g. "7:14PM M From notecards")
            if stripped_script and _STAGE_NOTE.match(stripped_script):
                continue

            # Skip garbled/fragmented text: if alpha words average < 3.5 chars it's
            # likely mis-detected column content (e.g. "e ecog Sayto te stage")
            if stripped_script:
                alpha_wds = [w for w in stripped_script.split() if w.isalpha()]
                if len(alpha_wds) >= 2:
                    avg_len = sum(len(w) for w in alpha_wds) / len(alpha_wds)
                    if avg_len < 3.5:
                        continue

            # Skip mid-sentence fragments and truncated page headers: if the first
            # alphabetic word starts lowercase it's a continuation, not a section title
            # (e.g. "screens.", "endure.", "ational Summit", "moments with us.")
            if stripped_script:
                alpha_words = [w for w in stripped_script.split() if w and w[0].isalpha()]
                if alpha_words and alpha_words[0][0].islower():
                    continue

            title = f"{ts_str}  {stripped_script}" if stripped_script else ts_str
            if len(title) > 65:
                title = title[:62] + "…"

            wi, li = _find_word_idx(pg_wds, first_script, used_wis)
            if wi is None:
                continue

            used_wis.add(wi)
            sections.append({
                "title":      title,
                "summary":    _make_summary(wi),
                "line_index": li,
                "word_index": wi,
            })

    if not sections:
        return None

    sections.sort(key=lambda s: s["word_index"])
    seen, unique = set(), []
    for s in sections:
        if s["word_index"] not in seen:
            seen.add(s["word_index"])
            unique.append(s)
    return unique


def _detect_script_column(pdf):
    """Detect the 'Script' column boundaries in a table-formatted PDF.

    Looks for column header row (Time | Script | Virtual Cues | ...) and uses
    the header x-positions to determine the Script column's left/right bounds.

    Returns (left_x, right_x) in PDF coordinates, or None if no table detected.
    Falls back gracefully so non-table PDFs are unaffected.
    """
    try:
        # Scan first few pages to find column headers
        for page_idx in range(min(len(pdf), 3)):
            page = pdf[page_idx]
            tp = page.get_textpage()
            n = tp.count_chars()
            if n == 0:
                continue

            limit = min(n, 3000)
            chars = []
            for i in range(limit):
                c = tp.get_text_range(i, 1)
                try:
                    box = tp.get_charbox(i)   # (left, bottom, right, top)
                    chars.append((c, box[0]))  # char, x_left
                except Exception:
                    chars.append((c, 0))

            full = "".join(c for c, _ in chars)

            # Find column headers — we need "Time" and "Virtual" at minimum
            headers = {}
            for hdr in ["Time", "Virtual", "Projection"]:
                # Search for each header, picking the occurrence with lowest x
                # (to avoid matching body text like "Time:" in a script line)
                idx = 0
                best_x = None
                while True:
                    found = full.find(hdr, idx)
                    if found < 0:
                        break
                    x = chars[found][1]
                    if best_x is None or x < best_x:
                        best_x = x
                        headers[hdr] = x
                    idx = found + len(hdr)

            if "Time" not in headers:
                continue

            # Script column: starts after "Time" header, ends at "Virtual" or "Projection"
            # Time column is narrow (~40-80px), Script body starts at ~80-100
            time_x = headers["Time"]
            # Script left = Time position + generous offset to clear the Time column
            script_left = time_x + 40

            # Script right = start of Virtual Cues or Projection column
            script_right = None
            for col in ["Virtual", "Projection"]:
                if col in headers:
                    script_right = headers[col]
                    break

            if script_right and script_right > script_left:
                print(f"[PDF] Detected Script column: x={script_left:.0f} → {script_right:.0f}")
                return (script_left, script_right)

        return None
    except Exception:
        return None


def _image_parse(content: bytes) -> dict:
    try:
        from PIL import Image
        import pytesseract
        text = pytesseract.image_to_string(Image.open(io.BytesIO(content)))
    except Exception:
        text = ""
    lines, words = [], []
    for raw in text.split("\n"):
        if raw.strip():
            _add_line(lines, words, raw)
    return _result(lines, words, None)  # HTML=None → <img> handles display


# ── Helpers ───────────────────────────────────────────────────────────────────

def _add_line(lines: list, words: list, text: str) -> int:
    """Append a line and its words; return the new line_index."""
    line_idx = len(lines)
    lines.append(text)
    for token in re.findall(r"\S+", text):
        words.append({
            "word": token,
            "clean": re.sub(r"[^\w']", "", token.lower()),
            "line_index": line_idx,
        })
    return line_idx


def _detect_sections(lines: list, words: list) -> list:
    """Detect meaningful navigation points for the operator.

    Detects in order of priority:
    1. Timestamp-prefixed lines (event rundowns): "650p Emcee: Welcome"
    2. Speaker cue lines (speaker scripts): "Kelly Russell:" alone on a line
    3. ALL CAPS phase headers: "START OF PROGRAM", "PANEL DISCUSSION"

    Each section includes a one-sentence summary extracted from its content.
    """
    line_to_word = {}
    for i, w in enumerate(words):
        li = w["line_index"]
        if li not in line_to_word:
            line_to_word[li] = i

    # Page header / AV-cue noise — skip entirely
    _NOISE = re.compile(
        r'^(Script\s*$|H\s*[–—\-]|for AV Team:|script to be|'
        r'Use your org|All user edits|Tech Notes|List Emcee|'
        r'Use Script Below|Record (Virtual|Below)|For Video Recording)',
        re.IGNORECASE,
    )
    _BULLET    = re.compile(r'^[●•○■◆]')   # actual bullet characters only
    _STAGE_DIR = re.compile(r'^[\[\(]')    # [pause] or (AV notes)         # [pause for applause]
    _TS = re.compile(
        r'^(\d{1,2}:\d{2}(?:\s*[ap]m)?|\d{3,4}[ap]|\d{1,2}[ap])\s+(.+)$',
        re.IGNORECASE,
    )
    # Speaker cue: optional "M " prefix, Name:, optional trailing (stage direction)
    # Matches: "Kelly Russell:"  "M Susan Lieu:"  "Sheryl Feldman: (in-person only)"
    _SPEAKER = re.compile(
        r'^(?:(?:PM?|EM|AM|MC)\s+)?(?:Dr\.\s+|Mr\.\s+|Ms\.\s+|Sen\.\s+)?'
        r'([A-Z][a-zA-Z\-]+(?: [A-Z][a-zA-Z\-]+){0,3}):\s*(?:\([^)]*\))?\s*$'
    )
    # Common non-person words that look like speaker names but aren't
    _NOT_SPEAKER = re.compile(
        r'^(Platform|Thermometer|Livestream|Donation|Website|Hashtag|'
        r'Registration|Parking|Venue|Location|Time|Date|Note|Script|Cue|'
        r'Audio|Video|Music|Slide|Screen|Camera|AV\s|AV$)\b',
        re.IGNORECASE,
    )

    def _is_caps_line(text: str) -> bool:
        a = re.sub(r"[^A-Za-z ]", "", text).strip()
        w = [x for x in a.split() if len(x) >= 2]
        return bool(a and a.upper() == a and len(w) >= 2)

    _SINGLE_LETTER_CUE = re.compile(r'^[A-Z]\s+')  # "M Something", "S Something"

    def _summary(start_li: int, end_li: int) -> str:
        """Return the first ~7 words of actual speech content after the section header."""
        for li in range(start_li + 1, min(end_li, start_li + 45)):
            text = lines[li].strip().rstrip('\r\n')
            if not text:
                continue
            if (_NOISE.match(text) or _STAGE_DIR.match(text)
                    or _is_caps_line(text) or _SINGLE_LETTER_CUE.match(text)):
                continue
            # Skip internal speaker cues — look past them to find speech text
            if _SPEAKER.match(text):
                continue
            # Strip bullet prefix to get the speech content
            text = re.sub(r'^[●•○■◆\-]\s*', '', text).strip()
            if not text or len(text) < 8:
                continue
            word_list = text.split()
            snippet = ' '.join(word_list[:7])
            if len(word_list) > 7:
                snippet += '…'
            return snippet
        return ""

    def _add(sections, li, title, end_li):
        title = title.strip()
        if len(title) > 65:
            title = title[:62] + "…"
        sections.append({
            "title": title,
            "summary": _summary(li, end_li),
            "line_index": li,
            "word_index": line_to_word[li],
        })

    # ── Pass 1: timestamp-prefixed lines (event rundowns) ────────────────────
    # Only count a line as a scheduled item if the content after the timestamp
    # starts with an uppercase letter — filters mid-sentence time references like
    # "8:30AM until then, this is team..." which are NOT agenda items.
    timestamped_lis = []
    for li, text in enumerate(lines):
        stripped = text.strip()
        if not stripped or li not in line_to_word:
            continue
        m = _TS.match(stripped)
        if m:
            name = m.group(2).strip()
            # Skip operational stage notes: "From notecards", "From script", "Notes:", etc.
            _ts_stage = re.compile(r'^(?:From\s+\w|Notes?:|TBD\b|N/A\b)', re.IGNORECASE)
            if name and name[0].isupper() and not _ts_stage.match(name):
                timestamped_lis.append((li, m.group(1), name))

    if timestamped_lis:
        sections = []
        for idx, (li, ts, name) in enumerate(timestamped_lis):
            end_li = timestamped_lis[idx + 1][0] if idx + 1 < len(timestamped_lis) else len(lines)
            _add(sections, li, f"{ts}  {name}", end_li)
        return sections

    # ── Pass 2: speaker-cue + ALL CAPS (speaker scripts) ────────────────────
    candidate_lis = []
    prev_was_caps = False   # track consecutive ALL CAPS lines to collapse multi-line headers

    for li, text in enumerate(lines):
        stripped = text.strip().rstrip('\r\n')
        if not stripped or li not in line_to_word:
            continue
        if _NOISE.match(stripped) or _BULLET.match(stripped) or _STAGE_DIR.match(stripped):
            prev_was_caps = False
            continue

        is_speaker = bool(_SPEAKER.match(stripped) and not _NOT_SPEAKER.match(stripped))

        is_caps = False
        if not is_speaker:
            alpha_only = re.sub(r"[^A-Za-z ]", "", stripped).strip()
            alpha_words = [w for w in alpha_only.split() if len(w) >= 2]
            if (alpha_only and alpha_only.upper() == alpha_only
                    and len(alpha_words) >= 2 and len(stripped.split()) <= 8
                    and not stripped.endswith(')')):  # skip truncated AV-note fragments
                is_caps = True

        if is_speaker:
            # Always add speaker cues — strip leading cue prefix (M, PM, EM, AM, MC)
            display = re.sub(r'^(?:PM?|EM|AM|MC)\s+', '', stripped)
            candidate_lis.append((li, display))
            prev_was_caps = False
        elif is_caps:
            if not prev_was_caps:
                # First line of an ALL CAPS block — add it
                candidate_lis.append((li, stripped))
            # Stay in caps mode regardless (continuation lines get skipped next iteration)
            prev_was_caps = True
        else:
            prev_was_caps = False

    # Deduplicate same title only within a 6-line window (back-to-back cue repeats)
    deduped = []
    for li, title in candidate_lis:
        if deduped and deduped[-1][1] == title and li - deduped[-1][0] < 6:
            continue
        deduped.append((li, title))

    sections = []
    for idx, (li, title) in enumerate(deduped):
        end_li = deduped[idx + 1][0] if idx + 1 < len(deduped) else len(lines)
        _add(sections, li, title, end_li)

    return sections


def _result(lines: list, words: list, html) -> dict:
    return {
        "lines": lines,
        "words": words,
        "word_count": len(words),
        "line_count": len(lines),
        "html": html,
        "sections": _detect_sections(lines, words),
    }


def _h(text: str) -> str:
    return _html.escape(str(text))


def _ocr_pdf(content: bytes) -> str:
    try:
        from pdf2image import convert_from_bytes
        images = convert_from_bytes(content)
        from PIL import Image
        import pytesseract
        return "\n".join(pytesseract.image_to_string(img) for img in images)
    except Exception:
        return ""
